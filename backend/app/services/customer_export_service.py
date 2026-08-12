from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.customer import Customer


ARCHIVE_HEADERS = (
    "客户名", "公司名", "国家", "获得客户时间", "职位", "WhatsApp", "邮箱", "电话",
    "来源", "客户类型", "兴趣产品", "客户体量", "客户等级", "客户总分", "跟进阶段",
    "自动判断阶段", "最近跟进日期", "是否需要跟进", "是否回复", "备注", "网站", "来源平台",
    "原始询盘内容", "CRM 客户评分", "CRM 销售阶段", "下次跟进日期", "最后跟进时间", "负责人",
)

TEXT_HEADERS = {"WhatsApp", "电话"}
DATE_HEADERS = {"获得客户时间", "最近跟进日期", "下次跟进日期"}
DATETIME_HEADERS = {"最后跟进时间"}
WRAPPED_HEADERS = {"备注", "原始询盘内容"}
WIDTHS = {
    "客户名": 18, "公司名": 28, "国家": 16, "获得客户时间": 14, "职位": 16,
    "WhatsApp": 20, "邮箱": 32, "电话": 20, "来源": 16, "客户类型": 16,
    "兴趣产品": 22, "客户体量": 12, "客户等级": 12, "客户总分": 12, "跟进阶段": 18,
    "自动判断阶段": 18, "最近跟进日期": 14, "是否需要跟进": 18, "是否回复": 12,
    "备注": 44, "网站": 32, "来源平台": 16, "原始询盘内容": 48, "CRM 客户评分": 14,
    "CRM 销售阶段": 18, "下次跟进日期": 14, "最后跟进时间": 20, "负责人": 18,
}


def _safe_cell_text(value: Any) -> str | None:
    """Keep archive text text and prevent spreadsheet formulas from executing."""
    if value is None:
        return None
    text = str(value)
    if text and text[0] in "=+-@":
        return f"'{text}"
    return text


def _stage_value(value: Any) -> str | None:
    if value is None:
        return None
    return _safe_cell_text(getattr(value, "value", value))


def _customer_values(customer: Customer) -> tuple[Any, ...]:
    return (
        customer.contact_name, customer.company_name, customer.country, customer.customer_acquired_at,
        customer.position, customer.whatsapp, customer.email, customer.phone, customer.source,
        customer.customer_type, customer.interested_product, customer.customer_size, customer.customer_level_value,
        customer.customer_total_score, customer.followup_stage, customer.automatic_stage_judgement,
        customer.latest_followup_date, customer.followup_requirement, customer.response_status, customer.notes,
        customer.website, customer.source_platform, customer.original_inquiry, customer.customer_score,
        _stage_value(customer.sales_stage), customer.next_followup_date, customer.last_followup_at,
        customer.owner.name if customer.owner is not None else None,
    )


def build_customer_archive_export(customers: list[Customer]) -> bytes:
    """Create a real UTF-8-safe .xlsx workbook for the complete customer archive."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "客户档案表"
    worksheet.freeze_panes = "A2"
    worksheet.append(ARCHIVE_HEADERS)
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(ARCHIVE_HEADERS))}1"

    header_fill = PatternFill("solid", fgColor="1E4B70")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for customer in customers:
        worksheet.append(_customer_values(customer))

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for index, cell in enumerate(row):
            header = ARCHIVE_HEADERS[index]
            if header in DATE_HEADERS and isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"
            elif header in DATETIME_HEADERS and isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm"
            elif header in TEXT_HEADERS:
                cell.value = _safe_cell_text(cell.value)
                cell.number_format = "@"
            elif isinstance(cell.value, str):
                cell.value = _safe_cell_text(cell.value)
            cell.alignment = Alignment(vertical="top", wrap_text=header in WRAPPED_HEADERS)

    for index, header in enumerate(ARCHIVE_HEADERS, 1):
        worksheet.column_dimensions[get_column_letter(index)].width = WIDTHS[header]

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
