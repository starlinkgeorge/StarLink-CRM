"""Create editable Excel workbooks from immutable quotation version snapshots."""

from io import BytesIO
import re

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import get_settings
from app.models.customer import Customer
from app.models.quotation import Quotation, QuotationVersion
from app.services.quotation_pdf_service import _fetch_public_image


BRAND_BLUE = "153A5B"
LIGHT_BLUE = "EAF1F7"
LIGHT_YELLOW = "FFF2CC"
LIGHT_GREY = "F8FAFC"
BORDER_GREY = "C9D4DF"
EDIT_NOTE = "Yellow cells are editable. Amounts recalculate automatically in Excel."


def quotation_excel_filename(quotation_number: str, version_no: int) -> str:
    """Return a download-safe Excel file name."""
    safe_number = re.sub(r"[^A-Za-z0-9_-]", "_", quotation_number)
    return f"{safe_number}-V{version_no}.xlsx"


def _currency_number_format(currency: str) -> str:
    safe_currency = re.sub(r"[^A-Za-z]", "", currency.upper())[:3] or "USD"
    return f'"{safe_currency}" #,##0.00'


def _add_picture(sheet, cell: str, image_url: str | None) -> None:  # noqa: ANN001
    """Embed a bounded image when its immutable snapshot can be resolved."""
    image_data = _fetch_public_image(image_url)
    if not image_data:
        sheet[cell] = "No image"
        sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
        return
    try:
        image = ExcelImage(BytesIO(image_data))
        image.width = 82
        image.height = 58
        sheet.add_image(image, cell)
    except Exception:
        # A malformed or unsupported image must not make the entire quotation
        # impossible to download.
        sheet[cell] = "No image"
        sheet[cell].alignment = Alignment(horizontal="center", vertical="center")


def generate_quotation_excel(
    quotation: Quotation,
    version: QuotationVersion,
    customer: Customer,
) -> bytes:
    """Build a user-editable business quotation workbook in memory."""
    settings = get_settings()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quotation"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A10"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:9"
    sheet.sheet_properties.outlinePr.summaryBelow = True

    for column, width in enumerate((36, 18, 15, 12, 18), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    thin_border = Side(style="thin", color=BORDER_GREY)
    table_border = Border(
        left=thin_border,
        right=thin_border,
        top=thin_border,
        bottom=thin_border,
    )
    brand_fill = PatternFill("solid", fgColor=BRAND_BLUE)
    soft_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    editable_fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
    soft_grey_fill = PatternFill("solid", fgColor=LIGHT_GREY)
    white_bold = Font(bold=True, color="FFFFFF")
    brand_bold = Font(bold=True, color=BRAND_BLUE)
    body_alignment = Alignment(vertical="center", wrap_text=True)
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    sheet.merge_cells("A1:E1")
    sheet["A1"] = "Dalian StarLink International Trade Co., Ltd."
    sheet["A1"].font = Font(bold=True, size=16, color=BRAND_BLUE)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells("A2:C4")
    sheet["A2"] = (
        f"Alibaba Store: {settings['company_alibaba_store'] or 'Not configured'}\n"
        f"Company Website: {settings['company_website'] or 'Not configured'}\n"
        f"Email: {settings['company_email'] or 'Not configured'}\n"
        f"WhatsApp: {settings['company_whatsapp'] or 'Not configured'}"
    )
    sheet["A2"].alignment = Alignment(vertical="top", wrap_text=True)
    sheet["A2"].font = Font(size=10, color="1F2937")

    sheet.merge_cells("D2:E4")
    sheet["D2"] = (
        "QUOTATION\n"
        f"Quotation No.: {quotation.quotation_number}\n"
        f"Version: V{version.version_no}\n"
        f"Date: {version.created_at:%Y-%m-%d}"
    )
    sheet["D2"].alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    sheet["D2"].font = Font(bold=True, color=BRAND_BLUE, size=11)
    for row in range(2, 5):
        sheet.row_dimensions[row].height = 20

    sheet.merge_cells("A6:B6")
    sheet.merge_cells("C6:E6")
    sheet["A6"] = "QUOTATION TO"
    sheet["C6"] = "CUSTOMER CONTACT"
    for cell in ("A6", "C6"):
        sheet[cell].fill = soft_fill
        sheet[cell].font = brand_bold
        sheet[cell].border = table_border
        sheet[cell].alignment = body_alignment
    for cell in ("B6", "D6", "E6"):
        sheet[cell].fill = soft_fill
        sheet[cell].border = table_border

    sheet.merge_cells("A7:B8")
    sheet.merge_cells("C7:E8")
    sheet["A7"] = f"{customer.company_name}\n{customer.country or '-'}"
    sheet["C7"] = (
        f"Contact: {customer.contact_name or '-'}\n"
        f"Email: {customer.email or '-'}\n"
        f"WhatsApp: {customer.whatsapp or '-'}"
    )
    for cell in ("A7", "C7"):
        sheet[cell].alignment = Alignment(vertical="top", wrap_text=True)
        sheet[cell].border = table_border
    for row in range(7, 9):
        sheet.row_dimensions[row].height = 24
        for column in range(1, 6):
            sheet.cell(row, column).border = table_border

    header_row = 9
    headers = ("Item Name", "Picture", "Unit Price", "QTY", "Total Price")
    for column, label in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column, label)
        cell.fill = brand_fill
        cell.font = white_bold
        cell.alignment = centered if column in {2, 4} else (
            right if column in {3, 5} else body_alignment
        )
        cell.border = table_border
    sheet.row_dimensions[header_row].height = 24

    item_start_row = header_row + 1
    number_format = _currency_number_format(version.currency)
    for index, item in enumerate(version.items):
        row = item_start_row + index
        name_cell = sheet.cell(row, 1, f"{item.product_name_snapshot}\n{item.sku_snapshot}")
        name_cell.font = Font(bold=True, color="1F2937")
        name_cell.alignment = body_alignment
        _add_picture(sheet, f"B{row}", item.picture_snapshot)
        unit_price = sheet.cell(row, 3, float(item.unit_price))
        quantity = sheet.cell(row, 4, float(item.quantity))
        line_total = sheet.cell(row, 5, f"=C{row}*D{row}")
        unit_price.fill = editable_fill
        quantity.fill = editable_fill
        unit_price.number_format = number_format
        quantity.number_format = "#,##0.00"
        line_total.number_format = number_format
        unit_price.alignment = right
        quantity.alignment = centered
        line_total.alignment = right
        for column in range(1, 6):
            sheet.cell(row, column).border = table_border
            if column not in {3, 4, 5}:
                sheet.cell(row, column).alignment = (
                    centered if column == 2 else body_alignment
                )
        sheet.row_dimensions[row].height = 62

    total_row = item_start_row + len(version.items) + 1
    for row, label, value in (
        (total_row, "Total cost", f"=SUM(E{item_start_row}:E{total_row - 2})"),
        (total_row + 1, "Door to door shipping cost", float(version.shipping_cost)),
        (total_row + 2, "Amount", f"=E{total_row}+E{total_row + 1}"),
    ):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        label_cell = sheet.cell(row, 1, label)
        value_cell = sheet.cell(row, 5, value)
        label_cell.alignment = body_alignment
        value_cell.alignment = right
        value_cell.number_format = number_format
        for column in range(1, 6):
            cell = sheet.cell(row, column)
            cell.border = table_border
        if row == total_row + 1:
            value_cell.fill = editable_fill
        if row == total_row + 2:
            for column in range(1, 6):
                cell = sheet.cell(row, column)
                cell.fill = brand_fill
                cell.font = white_bold
    sheet.cell(total_row, 5).font = Font(bold=True)

    terms_title_row = total_row + 5
    sheet.merge_cells(
        start_row=terms_title_row,
        start_column=1,
        end_row=terms_title_row,
        end_column=5,
    )
    sheet.cell(terms_title_row, 1, "TERMS AND CONDITIONS")
    sheet.cell(terms_title_row, 1).font = Font(bold=True, size=12, color=BRAND_BLUE)

    term_rows = (
        ("Validity (days)", version.validity_days, "0"),
        ("Payment Terms", version.payment_term, "@"),
        ("Delivery Time", version.delivery_time, "@"),
    )
    for offset, (label, value, number_format_code) in enumerate(term_rows, start=1):
        row = terms_title_row + offset
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        label_cell = sheet.cell(row, 1, label)
        value_cell = sheet.cell(row, 3, value)
        label_cell.fill = soft_fill
        label_cell.font = brand_bold
        value_cell.fill = editable_fill
        value_cell.number_format = number_format_code
        label_cell.alignment = body_alignment
        value_cell.alignment = body_alignment
        for column in range(1, 6):
            sheet.cell(row, column).border = table_border
        sheet.row_dimensions[row].height = 24

    note_row = terms_title_row + len(term_rows) + 2
    sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
    sheet.cell(note_row, 1, EDIT_NOTE)
    sheet.cell(note_row, 1).fill = soft_grey_fill
    sheet.cell(note_row, 1).font = Font(italic=True, color="4B5563", size=9)
    sheet.cell(note_row, 1).alignment = body_alignment
    sheet.row_dimensions[note_row].height = 22

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
