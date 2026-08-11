"""Import the ``客户档案表`` worksheet into the StarLink CRM safely.

The importer is deliberately idempotent.  It keeps existing customer IDs and
matches an existing record only by a persisted source-row key or an exact email,
WhatsApp number, or company/contact/country combination.  It never performs a
fuzzy merge.

Run a read-only workbook preflight:
    python scripts/import_customer_archive.py workbook.xlsx --validate-only

Run the database import after ``alembic upgrade head``:
    python scripts/import_customer_archive.py workbook.xlsx
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


SHEET_NAME = "客户档案表"
EXPECTED_HEADERS = (
    "获得客户时间",
    "来源",
    "客户名",
    "公司名",
    "职位",
    "备注",
    "国家",
    "客户类型",
    "兴趣产品",
    "WhatsApp",
    "邮箱",
    "电话",
    "客户等级",
    "客户体量",
    "客户总分",
    "跟进阶段",
    "自动阶段判断",
    "最近跟进日期",
    "是否回复",
    "是否需要跟进",
)

IDENTITY_HEADERS = ("客户名", "公司名", "WhatsApp", "邮箱", "电话")
DATE_HEADERS = {"获得客户时间", "最近跟进日期"}
INTEGER_HEADERS = {"客户等级", "客户体量", "客户总分"}


@dataclass(frozen=True)
class ArchiveRecord:
    row_number: int
    values: dict[str, Any]

    @property
    def company_name(self) -> str | None:
        return self.values["公司名"]

    @property
    def contact_name(self) -> str | None:
        return self.values["客户名"]

    @property
    def acquired_at(self) -> date | None:
        return self.values["获得客户时间"]


def clean_text(value: Any) -> str | None:
    """Return trimmed, UTF-8 safe text without converting identifiers to floats."""
    if value is None:
        return None
    if isinstance(value, str):
        text = value
    elif isinstance(value, bool):
        text = "是" if value else "否"
    elif isinstance(value, Decimal):
        text = format(value, "f")
    elif isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        if math.isnan(value):
            return None
        text = str(int(value)) if value.is_integer() else format(value, "f").rstrip("0").rstrip(".")
    else:
        text = str(value)
    # ``ignore`` removes invalid surrogate code points without changing valid
    # Chinese, English, accents, or other UTF-8 characters.
    text = text.encode("utf-8", errors="ignore").decode("utf-8").strip()
    return text or None


def parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        numeric = int(value)
        raw = str(numeric)
        if len(raw) == 8 and raw.isdigit():
            try:
                return datetime.strptime(raw, "%Y%m%d").date()
            except ValueError:
                return None
        try:
            converted = from_excel(float(value))
        except (TypeError, ValueError, OverflowError):
            return None
        if isinstance(converted, datetime):
            return converted.date()
        return converted if isinstance(converted, date) else None
    text = clean_text(value)
    if text is None:
        return None
    for pattern in ("%Y%m%d", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def parse_integer(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, Decimal)):
        return int(value)
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    text = clean_text(value)
    if text is None:
        return None
    try:
        numeric = Decimal(text)
    except Exception:
        return None
    return int(numeric) if numeric == numeric.to_integral_value() else None


def normalized(value: str | None) -> str | None:
    return value.casefold() if value else None


def composite_key(record: ArchiveRecord) -> tuple[str, str, str] | None:
    company = normalized(record.values["公司名"])
    contact = normalized(record.values["客户名"])
    country = normalized(record.values["国家"])
    return (company, contact, country) if company and contact and country else None


def row_is_effective(values: dict[str, Any]) -> bool:
    return any(values.get(header) is not None for header in IDENTITY_HEADERS)


def workbook_records(workbook_path: Path, sheet_name: str = SHEET_NAME) -> tuple[list[ArchiveRecord], dict[str, Any]]:
    """Read only the requested worksheet and return converted records/stats."""
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    formula_book = load_workbook(workbook_path, read_only=True, data_only=False)
    cached_book = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in formula_book.sheetnames:
            raise ValueError(f"Worksheet '{sheet_name}' was not found. Available: {formula_book.sheetnames}")
        formula_sheet = formula_book[sheet_name]
        cached_sheet = cached_book[sheet_name]
        header_row = next(formula_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = tuple(clean_text(value) for value in header_row)
        if headers != EXPECTED_HEADERS:
            missing = [header for header in EXPECTED_HEADERS if header not in headers]
            unexpected = [header for header in headers if header and header not in EXPECTED_HEADERS]
            raise ValueError(
                "客户档案表 columns changed. "
                f"Missing: {missing or 'none'}; unexpected: {unexpected or 'none'}"
            )

        records: list[ArchiveRecord] = []
        stats: dict[str, Any] = {
            "worksheet": sheet_name,
            "field_count": len(headers),
            "headers": list(headers),
            "sheet_data_rows": 0,
            "fully_blank_rows": 0,
            "formula_only_rows": 0,
            "invalid_rows": [],
        }
        formula_rows = formula_sheet.iter_rows(min_row=2)
        cached_rows = cached_sheet.iter_rows(min_row=2, values_only=True)
        for row_number, (formula_cells, cached_values) in enumerate(
            zip(formula_rows, cached_rows), start=2
        ):
            stats["sheet_data_rows"] += 1
            raw_formula_values = [cell.value for cell in formula_cells]
            if all(value is None or clean_text(value) is None for value in raw_formula_values):
                stats["fully_blank_rows"] += 1
                continue

            values: dict[str, Any] = {}
            for header, formula_cell, cached_value in zip(headers, formula_cells, cached_values):
                raw_value = cached_value if formula_cell.data_type == "f" else formula_cell.value
                if header in DATE_HEADERS:
                    values[header] = parse_date(raw_value)
                elif header in INTEGER_HEADERS:
                    values[header] = parse_integer(raw_value)
                else:
                    values[header] = clean_text(raw_value)

            if not row_is_effective(values):
                stats["formula_only_rows"] += 1
                continue
            if not values["公司名"]:
                stats["invalid_rows"].append(
                    {"row": row_number, "reason": "公司名为空，无法满足 CRM customers.company_name 必填约束"}
                )
                continue
            records.append(ArchiveRecord(row_number=row_number, values=values))
    finally:
        formula_book.close()
        cached_book.close()

    stats["valid_customers"] = len(records)
    return records, stats


class UnionFind:
    def __init__(self, count: int) -> None:
        self.parent = list(range(count))

    def find(self, index: int) -> int:
        while self.parent[index] != index:
            self.parent[index] = self.parent[self.parent[index]]
            index = self.parent[index]
        return index

    def union(self, first: int, second: int) -> None:
        first_root, second_root = self.find(first), self.find(second)
        if first_root != second_root:
            self.parent[second_root] = first_root


def select_unique_records(records: list[ArchiveRecord]) -> tuple[list[ArchiveRecord], list[dict[str, Any]]]:
    """Choose the most recent row from only exact, defensible duplicates."""
    union_find = UnionFind(len(records))
    seen_email: dict[str, int] = {}
    seen_whatsapp: dict[str, int] = {}
    seen_composite: dict[tuple[str, str, str], int] = {}

    for index, record in enumerate(records):
        for value, seen in (
            (normalized(record.values["邮箱"]), seen_email),
            (normalized(record.values["WhatsApp"]), seen_whatsapp),
            (composite_key(record), seen_composite),
        ):
            if value is None:
                continue
            if value in seen:
                union_find.union(index, seen[value])
            else:
                seen[value] = index

    grouped: dict[int, list[ArchiveRecord]] = {}
    for index, record in enumerate(records):
        grouped.setdefault(union_find.find(index), []).append(record)

    selected: list[ArchiveRecord] = []
    duplicates: list[dict[str, Any]] = []
    for group in grouped.values():
        winner = max(group, key=lambda item: (item.acquired_at or date.min, item.row_number))
        selected.append(winner)
        for item in group:
            if item is not winner:
                duplicates.append(
                    {
                        "duplicate_row": item.row_number,
                        "kept_row": winner.row_number,
                        "reason": "exact_email_or_whatsapp_or_company_contact_country",
                    }
                )
    return sorted(selected, key=lambda item: item.row_number), sorted(
        duplicates, key=lambda item: item["duplicate_row"]
    )


def archive_key(workbook_path: Path, sheet_name: str, row_number: int) -> str:
    return f"excel-customer-archive:{workbook_path.name}:{sheet_name}:{row_number}"


def customer_fields(record: ArchiveRecord) -> dict[str, Any]:
    """Map every workbook column to its customer table counterpart."""
    values = record.values
    return {
        "company_name": values["公司名"],
        "contact_name": values["客户名"],
        "country": values["国家"],
        "email": values["邮箱"],
        "phone": values["电话"],
        "whatsapp": values["WhatsApp"],
        "customer_acquired_at": values["获得客户时间"],
        "position": values["职位"],
        "notes": values["备注"],
        "customer_type": values["客户类型"],
        "source": values["来源"],
        "interested_product": values["兴趣产品"],
        "customer_level_value": values["客户等级"],
        "customer_size": values["客户体量"],
        "customer_total_score": values["客户总分"],
        "followup_stage": values["跟进阶段"],
        "automatic_stage_judgement": values["自动阶段判断"],
        "latest_followup_date": values["最近跟进日期"],
        "response_status": values["是否回复"],
        "followup_requirement": values["是否需要跟进"],
    }


def legacy_status_for(stage: str | None):
    """Return an existing CRM status only when the mapping is equivalent."""
    from app.models.customer import CustomerStatus

    return {
        "新开发未回复": CustomerStatus.LEAD,
        "新开发已回复": CustomerStatus.CONTACTED,
        "已报价": CustomerStatus.QUOTATION,
        "谈判中": CustomerStatus.NEGOTIATION,
        "已成交": CustomerStatus.WON,
        "已输单": CustomerStatus.LOST,
    }.get(stage)


def matching_customer(session, record: ArchiveRecord, source_key: str):
    """Find an existing CRM row using only exact identifiers."""
    from sqlalchemy import func, select

    from app.models.customer import Customer

    existing = session.scalar(select(Customer).where(Customer.archive_import_key == source_key))
    if existing is not None:
        return existing
    email = normalized(record.values["邮箱"])
    if email:
        existing = session.scalar(select(Customer).where(func.lower(Customer.email) == email))
        if existing is not None:
            return existing
    whatsapp = normalized(record.values["WhatsApp"])
    if whatsapp:
        existing = session.scalar(select(Customer).where(func.lower(Customer.whatsapp) == whatsapp))
        if existing is not None:
            return existing
    key = composite_key(record)
    if key:
        company, contact, country = key
        existing = session.scalar(
            select(Customer).where(
                func.lower(Customer.company_name) == company,
                func.lower(Customer.contact_name) == contact,
                func.lower(Customer.country) == country,
            )
        )
    return existing


def upsert_primary_contact(session, customer, record: ArchiveRecord) -> str | None:
    """Maintain one exact-name contact without inventing an extra person."""
    from sqlalchemy import func, select

    from app.models.customer import Contact

    contact_name = record.contact_name
    if not contact_name:
        return None
    contact = session.scalar(
        select(Contact).where(
            Contact.customer_id == customer.id,
            func.lower(Contact.name) == contact_name.casefold(),
        )
    )
    values = {
        "position": record.values["职位"],
        "email": record.values["邮箱"],
        "phone": record.values["电话"],
        "whatsapp": record.values["WhatsApp"],
    }
    if contact is None:
        session.add(Contact(customer_id=customer.id, name=contact_name, **values))
        return "created"
    for field, value in values.items():
        setattr(contact, field, value)
    return "updated"


def import_records(
    workbook_path: Path,
    records: Iterable[ArchiveRecord],
    sheet_name: str = SHEET_NAME,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Import records using a single transaction, reporting failures by Excel row."""
    from sqlalchemy import select

    backend_root = Path(__file__).resolve().parents[1]
    if str(backend_root) not in sys.path:
        sys.path.insert(0, str(backend_root))
    from app.db.session import get_session_factory
    from app.models.customer import Customer

    report: dict[str, Any] = {
        "new_customers": 0,
        "updated_customers": 0,
        "contacts_created": 0,
        "contacts_updated": 0,
        "failed_rows": [],
    }
    session = get_session_factory()()
    try:
        for record in records:
            try:
                with session.begin_nested():
                    source_key = archive_key(workbook_path, sheet_name, record.row_number)
                    customer = matching_customer(session, record, source_key)
                    created = customer is None
                    if customer is None:
                        customer = Customer(**customer_fields(record), archive_import_key=source_key)
                        legacy_status = legacy_status_for(record.values["跟进阶段"])
                        if legacy_status is not None:
                            customer.status = legacy_status
                            customer.sales_stage = legacy_status
                        session.add(customer)
                        session.flush()
                    else:
                        for field, value in customer_fields(record).items():
                            setattr(customer, field, value)
                        customer.archive_import_key = source_key
                        legacy_status = legacy_status_for(record.values["跟进阶段"])
                        if legacy_status is not None:
                            customer.status = legacy_status
                            customer.sales_stage = legacy_status
                        session.flush()

                    if record.values["最近跟进日期"] is not None:
                        customer.last_followup_at = datetime.combine(
                            record.values["最近跟进日期"], time.min, tzinfo=timezone.utc
                        )
                    result = upsert_primary_contact(session, customer, record)
                    report["new_customers" if created else "updated_customers"] += 1
                    if result:
                        report[f"contacts_{result}"] += 1
            except Exception as exc:  # one invalid row must not block valid customers
                report["failed_rows"].append({"row": record.row_number, "reason": str(exc)})
        if dry_run:
            session.rollback()
        else:
            session.commit()
        from sqlalchemy import func

        report["crm_customer_total"] = session.scalar(select(func.count(Customer.id)))
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()
    return report


def build_report(workbook_path: Path, sheet_name: str) -> tuple[list[ArchiveRecord], dict[str, Any]]:
    records, report = workbook_records(workbook_path, sheet_name)
    selected, duplicates = select_unique_records(records)
    report["selected_unique_customers"] = len(selected)
    report["explicit_duplicate_rows"] = len(duplicates)
    report["explicit_duplicates"] = duplicates
    return selected, report


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="Path to the uploaded .xlsx file")
    parser.add_argument("--sheet", default=SHEET_NAME, help="Worksheet name (default: 客户档案表)")
    parser.add_argument("--validate-only", action="store_true", help="Read and analyse without opening a DB")
    parser.add_argument("--dry-run", action="store_true", help="Check DB matching but rollback all changes")
    args = parser.parse_args()

    selected, report = build_report(args.workbook, args.sheet)
    if args.validate_only:
        report["mode"] = "validate_only"
        print(json.dumps(report, ensure_ascii=False, indent=2, default=str))
        return 0

    import_report = import_records(args.workbook, selected, args.sheet, dry_run=args.dry_run)
    report.update(import_report)
    report["mode"] = "dry_run" if args.dry_run else "imported"
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))
    return 0 if not report["failed_rows"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
