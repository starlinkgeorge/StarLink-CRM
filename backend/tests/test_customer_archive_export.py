from datetime import date, datetime, time, timezone
from io import BytesIO

from openpyxl import load_workbook

from app.models.customer import Customer
from app.services.customer_export_service import (
    ARCHIVE_HEADERS,
    build_customer_archive_export,
    normalize_excel_value,
)


def test_customer_archive_export_is_a_real_safe_xlsx() -> None:
    customer = Customer(
        company_name="Example School",
        contact_name="Amy",
        customer_acquired_at=date(2026, 8, 12),
        whatsapp="00123456789",
        phone="01234567890",
        notes="=This must remain text in Excel",
    )

    workbook = load_workbook(BytesIO(build_customer_archive_export([customer])))
    worksheet = workbook["客户档案表"]

    assert tuple(cell.value for cell in worksheet[1]) == ARCHIVE_HEADERS
    assert worksheet.auto_filter.ref == "A1:Z1"
    assert worksheet["A2"].value == "Amy"
    assert worksheet["D2"].value.date() == date(2026, 8, 12)
    assert worksheet["F2"].value == "00123456789"
    assert worksheet["F2"].number_format == "@"
    assert worksheet["H2"].value == "01234567890"
    assert worksheet["H2"].number_format == "@"
    assert worksheet["R2"].value == "'=This must remain text in Excel"


def test_customer_archive_export_normalizes_timezone_aware_values() -> None:
    utc_timestamp = datetime(2026, 8, 12, 18, 45, tzinfo=timezone.utc)
    customer = Customer(
        company_name="Timezone School",
        contact_name="Timezone Buyer",
        customer_acquired_at=utc_timestamp,
        latest_followup_date=utc_timestamp,
        last_followup_at=utc_timestamp,
        whatsapp="00123456789",
        phone="01234567890",
    )

    workbook = load_workbook(BytesIO(build_customer_archive_export([customer])))
    worksheet = workbook["客户档案表"]

    # UTC 18:45 is 02:45 on the following China business date.
    assert worksheet["D2"].value.date() == date(2026, 8, 13)
    assert worksheet["Q2"].value.date() == date(2026, 8, 13)
    assert worksheet["Y2"].value == datetime(2026, 8, 13, 2, 45)
    assert worksheet["Y2"].value.tzinfo is None
    assert worksheet["F2"].value == "00123456789"
    assert worksheet["F2"].number_format == "@"
    assert worksheet["H2"].value == "01234567890"
    assert worksheet["H2"].number_format == "@"
    assert normalize_excel_value(time(10, 30, tzinfo=timezone.utc)).tzinfo is None
